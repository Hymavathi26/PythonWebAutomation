
import os
from dotenv import load_dotenv
from selenium import webdriver
from selenium.webdriver.common.by import By
import time
import pytest
import openpyxl
from openpyxl import load_workbook



def read_the_credential_from_excel(file_path):
    # how to read file by using openpyxl
    credentials = []
    workbook = openpyxl.load_workbook(file_path)  # take excel file path
    sheet = workbook.active  # take active excel sheet

    for row in sheet.iter_rows(values_only=True):  #row start from 2 row in and values_only not formatting just given nomal values
        username, password = row
        credentials.append({"username": username, "password": password})
        return credentials

@pytest.mark.parametrize("user_cred",read_the_credential_from_excel("/Users/HymavathiKothapalle/pythonProject/PythonWebAutomation/code/12_jandatadrivenClass/test_betterway_TC1.py"))
def test_vwo_login_positive(user_cred):
    username = user_cred["username"]
    password = user_cred["password"]
    vwo_login(username,password)



def vwo_login(username,password):
    driver = webdriver.Chrome()
    driver.get("https://app.vwo.com")
    email_input = driver.find_element(By.XPATH, "//input[@id='login-username']")
    pw_input = driver.find_element(By.XPATH, "//input[@id='login-password']")
    submit_input = driver.find_element(By.CSS_SELECTOR, "#js-login-btn")

    email_input.send_keys(username)
    pw_input.send_keys(password)
    submit_input.click()

    result=driver.current_url
    if result != "https://app.vwo.com/#dashboard":
        pytest.xfail("Invalid login")
    driver.quit()

    #write the logic if tstcase pass
    #-->URL changes TC
    #-->Error message
